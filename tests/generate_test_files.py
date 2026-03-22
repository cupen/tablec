#!/usr/bin/env python3
"""Generate test Excel files for tablec testing."""

import openpyxl
from openpyxl.styles import Font
from pathlib import Path

def create_basic_types_test():
    """Create Excel file with basic types test."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BasicTypes"

    # Header rows (first 5 reserved rows)
    headers = [
        ["id", "name", "score", "active", "price", "description"],  # Row 1: Field names
        ["int", "str", "float64", "bool", "float32", "string"],  # Row 2: Field types
        ["ID", "Name", "Score", "Active", "Price", "Description"],  # Row 3: Comments
        ["@unique", "", "", "", "", ""],  # Row 4: Constraints
        ["", "", "", "", "", ""]  # Row 5: Reserved
    ]

    for i, row in enumerate(headers, 1):
        for j, cell in enumerate(row, 1):
            ws.cell(row=i, column=j, value=cell)
            if i <= 3:
                ws.cell(row=i, column=j).font = Font(bold=True)

    # Data rows (starting from row 6)
    data = [
        [1, "Alice", 95.5, True, 10.99, "First user"],
        [2, "Bob", 87.3, True, 25.50, "Second user"],
        [3, "Charlie", 92.1, False, 15.75, "Third user"],
        [4, "Diana", 88.9, True, 8.99, "Fourth user"],
        [5, "Eve", 94.7, False, 12.25, "Fifth user"]
    ]

    for i, row in enumerate(data, 6):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)

    output_path = Path(__file__).parent / "excel" / "basic_types.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return wb

def create_array_types_test():
    """Create Excel file with array types test."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ArrayTypes"

    headers = [
        ["id", "tags", "scores", "matrix"],
        ["int", "string[]", "int64[]", "float32[][]"],
        ["ID", "Tags", "Scores", "Matrix"],
        ["@unique", "", "", ""],
        [""]
    ]

    for i, row in enumerate(headers, 1):
        for j, cell in enumerate(row, 1):
            ws.cell(row=i, column=j, value=cell)
            if i <= 3:
                ws.cell(row=i, column=j).font = Font(bold=True)

    data = [
        [1, "['tag1', 'tag2']", "[100, 95, 90]", "[[1.0, 2.0], [3.0, 4.0]]"],
        [2, "['tag3', 'tag4', 'tag5']", "[88, 92, 87, 93]", "[[5.0, 6.0], [7.0, 8.0]]"],
        [3, "['tag1']", "[95]", "[[9.0, 10.0]]"],
        [4, "[]", "[85, 80]", "[[1.1, 2.2], [3.3, 4.4], [5.5, 6.6]]"]
    ]

    for i, row in enumerate(data, 6):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)

    output_path = Path(__file__).parent / "excel" / "array_types.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return wb

def create_map_types_test():
    """Create Excel file with map types test."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MapTypes"

    headers = [
        ["id", "properties", "int_map"],
        ["int", "map<string, string>", "map<int, int64>"],
        ["ID", "Properties", "IntMap"],
        ["@unique", "", ""],
        [""]
    ]

    for i, row in enumerate(headers, 1):
        for j, cell in enumerate(row, 1):
            ws.cell(row=i, column=j, value=cell)
            if i <= 3:
                ws.cell(row=i, column=j).font = Font(bold=True)

    data = [
        [1, "{'key1': 'value1', 'key2': 'value2'}", "{1: 10, 2: 20}"],
        [2, "{'name': 'Alice', 'role': 'admin'}", "{1: 5, 3: 15, 5: 25}"],
        [3, "{}", "{10: 100}"]
    ]

    for i, row in enumerate(data, 6):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)

    output_path = Path(__file__).parent / "excel" / "map_types.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return wb

def create_struct_types_test():
    """Create Excel file with struct types test."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StructTypes"

    headers = [
        ["id", "position", "user_info"],
        ["int", "struct{x: int, y: int}", "struct{name: str, age: int32, active: bool}"],
        ["ID", "Position", "UserInfo"],
        ["@unique", "", ""],
        [""]
    ]

    for i, row in enumerate(headers, 1):
        for j, cell in enumerate(row, 1):
            ws.cell(row=i, column=j, value=cell)
            if i <= 3:
                ws.cell(row=i, column=j).font = Font(bold=True)

    data = [
        [1, "{10, 20}", "{Alice, 25, True}"],
        [2, "{30, 40}", "{Bob, 30, False}"],
        [3, "{50, 60}", "{Charlie, 35, True}"]
    ]

    for i, row in enumerate(data, 6):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)

    output_path = Path(__file__).parent / "excel" / "struct_types.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return wb

def create_constraints_test():
    """Create Excel file with constraint tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Constraints"

    headers = [
        ["id", "name", "seq_num", "score"],
        ["int", "string", "int64", "float32"],
        ["ID", "Name", "Sequence", "Score"],
        ["@unique", "", "@seq", "@order(asc)"],
        [""]
    ]

    for i, row in enumerate(headers, 1):
        for j, cell in enumerate(row, 1):
            ws.cell(row=i, column=j, value=cell)
            if i <= 3:
                ws.cell(row=i, column=j).font = Font(bold=True)

    data = [
        [1, "Alice", 1, 85.5],
        [2, "Bob", 2, 87.3],
        [3, "Charlie", 3, 88.9],
        [4, "Diana", 4, 90.1],
        [5, "Eve", 5, 92.5]
    ]

    for i, row in enumerate(data, 6):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)

    output_path = Path(__file__).parent / "excel" / "constraints.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return wb

def create_composite_types_test():
    """Create Excel file with composite/nested types test."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CompositeTypes"

    headers = [
        ["id", "points", "metadata"],
        ["int", "struct{x: int, y: int}[]", "map<string, int[]>"],
        ["ID", "Points", "Metadata"],
        ["@unique", "", ""],
        [""]
    ]

    for i, row in enumerate(headers, 1):
        for j, cell in enumerate(row, 1):
            ws.cell(row=i, column=j, value=cell)
            if i <= 3:
                ws.cell(row=i, column=j).font = Font(bold=True)

    data = [
        [1, "[{10, 20}, {30, 40}]", "{'tags': [1, 2, 3], 'ids': [10, 20]}"],
        [2, "[{0, 0}, {100, 100}]", "{'levels': [1, 2, 3, 4]}"]
    ]

    for i, row in enumerate(data, 6):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)

    output_path = Path(__file__).parent / "excel" / "composite_types.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return wb

if __name__ == "__main__":
    print("Generating test Excel files...")
    create_basic_types_test()
    create_array_types_test()
    create_map_types_test()
    create_struct_types_test()
    create_constraints_test()
    create_composite_types_test()
    print("\nAll test Excel files generated successfully!")
